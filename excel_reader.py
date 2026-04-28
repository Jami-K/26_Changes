import hashlib
import openpyxl

from database import upsert_change

# 엑셀 헤더명 → DB 컬럼명 매핑
COLUMN_MAP = {
    '고지일':       'notice_date',
    '제품코드':     'product_code',
    '제품명':       'product_name',
    '변경자재':     'changed_material',
    '혼용가능여부': 'mixable',
    '변경내용':     'change_content',
    '적용예정일':   'apply_date',
    '특이사항':     'notes',
    '디자인파일경로': 'design_file',
}


def _str(v):
    if v is None:
        return ''
    s = str(v).strip()
    # 날짜가 20260428 형태의 8자리 정수로 저장된 경우 YYYY-MM-DD 로 변환
    if s.isdigit() and len(s) == 8:
        s = f'{s[:4]}-{s[4:6]}-{s[6:]}'
    return s


def sync_excel(path: str) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [_str(c.value) for c in next(ws.iter_rows(max_row=1))]

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue

        data = {v: '' for v in COLUMN_MAP.values()}
        for i, header in enumerate(headers):
            if header in COLUMN_MAP and i < len(row):
                data[COLUMN_MAP[header]] = _str(row[i])

        hash_src = '|'.join([
            data['notice_date'], data['product_code'],
            data['product_name'], data['changed_material'],
            data['apply_date'],
        ])
        data['row_hash'] = hashlib.sha256(hash_src.encode('utf-8')).hexdigest()[:32]

        inserted += upsert_change(data)

    wb.close()
    return inserted
